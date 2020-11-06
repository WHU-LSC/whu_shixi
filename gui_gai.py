#-*- coding: utf-8 -*-
'''
Created on Wed Jan  8 21:15:13 2020

@author: Administrator
'''

#爬虫的交互界面编程
from selenium import webdriver
import re
import sys
import os
import time
from  openpyxl import Workbook
from openpyxl.styles import Font,Alignment
import requests
from bs4 import BeautifulSoup
import datetime
import urllib
from fontTools.ttLib import TTFont
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By 


#主界面——登录
class main():
    def __init__(self):
        self.main()
    #—————————————————————————————————————可能会用到的函数，摆在前面—————————————————————————————————————
    #用于list的嵌套展开
    def flatten(self,input_list):
        output_list = []
        while True:
            if input_list == []:
                break
            for index, i in enumerate(input_list):
    
                if type(i) == list:
                    input_list = i + input_list[index + 1:]
                    break
                else:
                    output_list.append(i)
                    input_list.pop(index)
                    break
        return output_list
    #整合数据用的函数
    def zhenghe(self,a,b):#用于数据整合，避免使用numpy
        re=[]
        for x in a:
            re.append(x)
        for x in b:
            re.append(x)
        return re  
    #—————————————————————————————————————宣讲会信息的获取框架—————————————————————————————————————
    #用于单日的数据写入
    def write_Excel_jiuye(self,data,row_now,wb):
        d_now=datetime.datetime.now()
        d_now=d_now+datetime.timedelta(1)
        d_next=d_now+datetime.timedelta(1)
        sheet = wb.active
        sheet.title =str(d_now.strftime('%m'))+'月'+str(d_now.strftime('%d'))+'日-'+str(d_next.strftime('%m'))+'月'+str(d_next.strftime('%d'))+'日'+'就业宣讲会信息汇总' 
        sheet.cell(row=row_now, column=1, value='地区')
        sheet.cell(row=row_now, column=2, value='宣讲会名称')
        sheet.cell(row=row_now, column=3, value='地点')
        sheet.cell(row=row_now, column=4, value='时间')
        sheet.cell(row=row_now, column=5, value='链接')
        for i in range(5):
            sheet.cell(row=row_now, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
        #写入
        for i in range(len(data)):
            for j in range(1, len(data[i])+1):
                if(j == 4):
                    sheet.cell(row=i +row_now+ 1, column=5, value=('=HYPERLINK("%s","%s")' % (data[i][3], '详情')))
                    sheet.cell(row=i +row_now+ 1, column=5).alignment = Alignment(horizontal='center', vertical='center')
                else:
                    sheet.cell(row=i +row_now+ 1, column=j+1, value=str(data[i][j-1]))
        #保存
        try:
            wb.save(r"D:/发展权益部/明后就业宣讲会信息.xlsx")
            #print('文件保存在: D:/发展权益部/明后就业宣讲会信息.xls')
        except FileNotFoundError:
            try:
                os.makedirs(r"D:/发展权益部")
                wb.save(r"D:/发展权益部/明后就业宣讲会信息.xlsx")
                #print('文件保存在: D:/发展权益部/明后就业宣讲会信息.xls')
            except FileNotFoundError:
                os.makedirs(r"C:/发展权益部")
                wb.save(r"C:/发展权益部/明后就业宣讲会信息.xlsx")
                #print('文件保存在: C:/发展权益部/明后就业宣讲会信息.xls')
    #用于多日文件写入——wrotr——main
    def wirte_jiuye_total(self,result):
        #拆分数据
        wb = Workbook()
        re_now=[]
        re_next=[]
        d_now=datetime.datetime.now()
        d_now=d_now+datetime.timedelta(1)
        d_next=d_now+datetime.timedelta(1)
        for x in result:
            if x[2][5:10]==d_now.strftime('%m-%d'):
                re_now.append(x)
            elif x[2][5:10]==d_next.strftime('%m-%d'):
                re_next.append(x)
        wb = Workbook()
        sheet = wb.active
        #sheet.title = str(d_now.strftime('%m-%d'))+'和'+str(d_next.strftime('%m-%d'))+'就业宣讲会信息汇总' # 创建最终保存表格
        sheet.title =str(d_now.strftime('%m'))+'月'+str(d_now.strftime('%d'))+'日-'+str(d_next.strftime('%m'))+'月'+str(d_next.strftime('%d'))+'日'+'就业宣讲会信息汇总'
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 30
        #第一天
        sheet.cell(row=1, column=1, value=str(d_now.strftime('%m'))+'月'+str(d_now.strftime('%d'))+'日'+'就业宣讲会信息')
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        self.write_Excel_jiuye(re_now,2,wb)
        #第二天
        sheet.cell(row=len(re_now)+3, column=1, value=str(str(d_next.strftime('%m'))+'月'+str(d_next.strftime('%d'))+'日'+'就业宣讲会信息'))
        sheet.cell(row=len(re_now)+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=len(re_now)+3, start_column=1, end_row=len(re_now)+3, end_column=5)
        sheet.cell(row=len(re_now)+len(re_next)+5, column=1, value='武大研会发展权益部')
        sheet.merge_cells('A{}:E{}'.format(len(re_now)+len(re_next)+5,len(re_now)+len(re_next)+5))
        sheet['A{}'.format(len(re_now)+len(re_next)+5)].alignment = Alignment(horizontal='center', vertical='center')
        self.write_Excel_jiuye(re_next,len(re_now)+4,wb)
    #获取就业宣讲会信息并整合
    def shuju_jiuye(self,browser):
        #获取宣讲会名称mc
        mc=[]
        mc_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="tabGrid"]/tbody/tr/td[6]')
        for x in pwd:
            mc_test=[x.text,x.get_attribute('href')]
            mc.append(mc_test[0])    
            #获取宣讲会地点dd
        dd=[]
        dd_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="tabGrid"]/tbody/tr/td[8]')
        for x in pwd:
            dd_test=[x.text,x.get_attribute('href')]
            dd.append('武汉大学'+str(dd_test[0]))
        
        #获取宣讲会时间sj
        time = []
        sj_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="tabGrid"]/tbody/tr/td[9]')
        for x in pwd:
            sj_test=[x.text,x.get_attribute('href')]
            time.append(sj_test[0])
        #获取详情xq
        xq = []
        xq_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="tabGrid"]/tbody/tr/td[6]')
        for x in pwd:
            xq_test=[x.get_attribute('innerHTML')]
            xq.append('http://www.xsjy.whu.edu.cn/zftal-web/zfjy!wzxx/'+str(re.findall(r'&quot;(.*?);',str(xq_test[0]))[1]))
        #要的信息全部整理完毕，进行转置获得需要的格式
        return [[mc[i],dd[i],time[i],xq[i]] for i in range(len(mc))]
    #就业宣讲会信息获取主函数
    def jiuye_main(self):
        print('开始抓取宣讲会信息......\n')
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')
        #将当前目录放到搜索栏当中
        path = os.path.abspath(os.curdir)
        sys.path.append(path)
        browser = webdriver.Chrome(r'D:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe', options=chrome_options)
        #这里通过get请求需要模拟登录的页面
        browser.get("http://www.xsjy.whu.edu.cn/zftal-web/zfjy!wzxx/xjhxx_cxXjhForWeb.html")
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        #browser.manage().timeouts().implicitlyWait(2000, TimeUnit.SECONDS);#延迟
        k = 3
        if k==1:
            result = self.shuju_jiuye(browser)
        else: 
            result = self.shuju_jiuye(browser)
            for i in range(k-1):
                button = WebDriverWait(browser,10).until(EC.element_to_be_clickable((By.LINK_TEXT,"下一页")))
                button.click()  
                result += self.shuju_jiuye(browser)
        if result ==[]:
            print('网站似乎没有更新数据，可能是明后两天没有宣讲会......\n')
        else:
            print('信息已经检索完毕，现在开始写入......\n')
            self.wirte_jiuye_total(result)
            print('信息已经写入完毕!\n')  
            print('===============\n') 
        browser.quit()
      
       
    #—————————————————————————————————————实习信息的获取框架—————————————————————————————————————
  
    #接下来写入文件（包括全国地区的内容和武汉地区的内容）
    def write_Excel_shixi(self,data,result_shixisen_wuhan):
        wb = Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 9.0
        sheet.column_dimensions['B'].width = 50.0
        sheet.column_dimensions['C'].width = 50.0
        sheet.column_dimensions['D'].width = 13.0
        sheet.title = datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"实习信息" 
        sheet.cell(row=1, column=1, value=datetime.datetime.now().strftime('%Y')+'年'+datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"实习信息")
        sheet.merge_cells('A1:D1')
        #sheet.cellstyle('A1', font, align)
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # 创建最终保存表格
        sheet.cell(row=2, column=1, value='地区')
        sheet.cell(row=2, column=2, value='公司名称')
        sheet.cell(row=2, column=3, value='岗位')
        sheet.cell(row=2, column=4, value='链接')
        sheet.cell(row=2, column=5, value='来源')
        for i in range(5):
            for j in range(2,len(data)+3):
                sheet.cell(row=j, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
        
        for i in range(len(data)):
            for j in range(len(data[i])):
                if(j == 3):
                    sheet.cell(row=i + 3, column=j+1, value=('=HYPERLINK("%s","%s")' % (data[i][j], '详情')))
                else:
                    sheet.cell(row=i + 3, column=j+1, value=str(data[i][j]))
        #sheet['A3:D{}'.format(len(data)+2)].alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=len(data)+3, column=1, value='武大研会发展权益部')
        sheet.merge_cells('A{}:D{}'.format(len(data)+3,len(data)+3))
        sheet['A{}'.format(len(data)+3)].alignment = Alignment(horizontal='center', vertical='center')
        #开启第二个sheet
        if result_shixisen_wuhan == []:
            pass
        else:
            #首先要去除重复
            #result_shixisen_wuhan = list(set(result_shixisen_wuhan))
            list2=[]
            for i in result_shixisen_wuhan:
                if i not in list2:
                    list2.append(i)
            result_shixisen_wuhan=list2
            sheet1 = wb.create_sheet(datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"武汉地区实习信息" )
            sheet1.column_dimensions['A'].width = 9.0
            sheet1.column_dimensions['B'].width = 50.0
            sheet1.column_dimensions['C'].width = 50.0
            sheet1.column_dimensions['D'].width = 13.0
            #sheet1.title = datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"实习信息" 
            sheet1.cell(row=1, column=1, value=datetime.datetime.now().strftime('%Y')+'年'+datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"武汉地区实习信息")
            sheet1.merge_cells('A1:D1')
            #sheet.cellstyle('A1', font, align)
            sheet1['A1'].alignment = Alignment(horizontal='center', vertical='center')
            # 创建最终保存表格
            sheet1.cell(row=2, column=1, value='地区')
            sheet1.cell(row=2, column=2, value='公司名称')
            sheet1.cell(row=2, column=3, value='岗位')
            sheet1.cell(row=2, column=4, value='链接')
            sheet1.cell(row=2, column=5, value='来源')
            for i in range(5):
                for j in range(2,len(result_shixisen_wuhan)+3):
                    sheet1.cell(row=j, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
            
            for i in range(len(result_shixisen_wuhan)):
                for j in range(len(result_shixisen_wuhan[i])):
                    if(j == 3):
                        sheet1.cell(row=i + 3, column=j+1, value=('=HYPERLINK("%s","%s")' % (result_shixisen_wuhan[i][j], '详情')))
                    else:
                        sheet1.cell(row=i + 3, column=j+1, value=str(result_shixisen_wuhan[i][j]))
            #sheet['A3:D{}'.format(len(data)+2)].alignment = Alignment(horizontal='center', vertical='center')
            sheet1.cell(row=len(result_shixisen_wuhan)+3, column=1, value='武大研会发展权益部')
            sheet1.merge_cells('A{}:D{}'.format(len(result_shixisen_wuhan)+3,len(result_shixisen_wuhan)+3))
            sheet1['A{}'.format(len(result_shixisen_wuhan)+3)].alignment = Alignment(horizontal='center', vertical='center')
        #保存
        try:
            wb.save(r"D:/发展权益部/{}实习信息.xlsx".format(datetime.datetime.now().strftime('%m-%d')))
        except FileNotFoundError:
            try:
                os.makedirs(r"D:/发展权益部")
                wb.save(r"D:/发展权益部/{}实习信息.xlsx".format(datetime.datetime.now().strftime('%m-%d')))
            except FileNotFoundError:
                os.makedirs(r"C:/发展权益部")
                wb.save(r"C:/发展权益部/{}实习信息.xlsx".format(datetime.datetime.now().strftime('%m-%d')))
    # 应届生求职网的主函数
    def get_info_yingjiesheng(self):
        print('开始抓取实习信息......\n')
        #加载启动项
        try:
            option = webdriver.ChromeOptions()
            option.add_argument('headless')
            #import win32com
            #应届生求职网的信息
            url = 'http://www.yingjiesheng.com/commend-parttime-1.html'
            wb_data = requests.get(url)
            soup = BeautifulSoup(wb_data.content, 'html.parser', from_encoding='utf-8')
            a = []
            time=[]
            result = []
            for link in soup.select('#mainNav > div.jobList > table > tr > td > a'):
                                    #mainNav > div.jobList > table > tbody > tr:nth-child(22) > td.date
                a.append(str(link))
            for link in soup.select('#mainNav > div.jobList > table > tr > td.date'):
                                    #mainNav > div.jobList > table > tbody > tr:nth-child(22) > td.date
                time.append(str(link))
            
            for i in range(len(a)):
                result.append([re.findall(r'<a href="(.*?)" target=',a[i]),
                               re.findall(r'#008000;">(.*?)</span>',a[i]),
                               re.findall(r'</span>(.*?)</a>',a[i]),
                               re.findall(r'<td class="date">(.*?)</td>',time[i])[0]])
            for i in range(len(result)):
                if result[i][-1][-5:]!=datetime.datetime.now().strftime('%m-%d'):
                    result[i]=[]
            result=[x[1]+x[2]+x[0] for x in result if x!=[]]
            for i in range(len(result)):
                if len(result[i])!=3:
                    result[i]=[]
            result=[x for x in result if x!=[]]
            for x in result:
                x[1]=x[1].split("招聘")
                if len(x[1])==1:
                    x[1].append('实习生')
                elif len(x[1])!=2:
                    del x[1][1]
                    if x[1][1]=='':
                        x[1][1]='实习生'
                elif x[1][1]=='':
                    x[1][1]='实习生'
                elif len(x[1])==1:
                    x[1].append('实习生')
            result = [self.flatten(x) for x in result if x!=[]]
            for x in result:
                x[1]=x[1].replace('实习生','')
                x[1]=x[1].replace('2020','')
                x[1]=x[1].replace('春季','')
                x[1]=x[1].replace('秋季','')
                x[1]=x[1].replace('暑假','')
                x[1]=x[1].replace('寒假','')
                x[1]=x[1].replace('暑期','')
                x[1]=x[1].replace('冬季','')
                if '<' in x[1] and '>' in x[1]:
                    x[1]=x[1].split('</span>')[1]
            for i in range(len(result)):
                    if 'http' not in result[i][-1]:
                        result[i][-1]='http://www.yingjiesheng.com/'+result[i][-1]
            for x in result:
                x.append('应届生求职网')
            return result    
        except IndexError:
            return []
    # 应届生求职网的备用函数
    def getinfo_new4(self,url):
        a=[]
        wb_data = requests.get(url)
        url=[]
        soup = BeautifulSoup(wb_data.content, 'html.parser', from_encoding='utf-8')
        try:
            for link in soup.select('#wrap > div.clear > div.rec.recr > ul'):
                                    ##wrap > div.clear > div.rec.recr > ul > li:nth-child(1)
                a.append(link)
                url.append(str(link))
            #print(len(a),len(str(a[0]).split('实习生'))>=4)
            #print(a)
            if len(str(a[0]).split('实习生'))<=4 :
                a=[]
                url=[]
                for link in soup.select('#mainNav > div.recommend.s_clear > div.box.floatr > ul:nth-child(8)'):
                    a.append(link)
                    url.append(str(link))
            if len(str(a[0]).split('实习生'))<=4:
                a=[]
                url=[]
                for link in soup.select('#mainNav > div.recommend.s_clear > div.box.floatr > ul:nth-child(6)'):
                    a.append(link)
                    url.append(str(link))
            a = str(a[0]).split('</li>\n<li>\n')
            a = self.flatten([x.split('li>\n<li') for x in a])
            a=[x for x in a if '.'+datetime.datetime.now().strftime('%d') in x]
               #datetime.datetime.now().strftime('%d')
    #         #url=[re.findall(r'href="(.*?)" target=',str(x))[0] for x in a if x!='']
            a=[[re.findall(r'#008000;">(.*?)</span>',str(x)),re.findall(r'</span>(.*?)</a>',str(x)),re.findall(r'href="(.*?)" target=',str(x))] for x in a] 
            info = [self.flatten(a[i]) for i in range(len(a))]
            for i in range(len(info)):
                if len(info[i])<3:
                    info[i]=[]
            info = [x for x in info if x!=[]]
            for i in range(len(info)):
                    if 'http' not in info[i][-1]:
                        info[i][-1]='http://www.yingjiesheng.com/'+info[i][-1]
    
            for x in info:
                x[1]=x[1].split("招聘")
                if len(x[1])==1:
                    x[1].append('实习生')
                elif len(x[1])!=2:
                    del x[1][1]
                    if x[1][1]=='':
                        x[1][1]='实习生'
                elif x[1][1]=='':
                    x[1][1]='实习生'
                elif len(x[1])==1:
                    x[1].append('实习生')
            info = [self.flatten(x) for x in info if x!=[]]
            info = [x[:3]+[x[-1],'应届生'] for x in info if x!=[]]
            for x in info:
                x[1]=x[1].replace('实习生','')
                x[1]=x[1].replace('2020','')
                x[1]=x[1].replace('2021','')
                x[1]=x[1].replace('春季','')
                x[1]=x[1].replace('秋季','')
                x[1]=x[1].replace('暑假','')
                x[1]=x[1].replace('寒假','')
                x[1]=x[1].replace('暑期','')
                x[1]=x[1].replace('冬季','')
            return info
        except:
            return []

    #获取实习僧信息的爬虫函数
    def shuju_shixisen(self,page,browser,url):
        if url !='no_url':
            browser.get(url.format(page))
        if page==1:
            ttf = []
            ttf_test=[]
            pwd = browser.find_elements_by_xpath('/html/head/style[1]')
            for x in pwd:
                ttf_test=[x.get_attribute('outerHTML')]
                ttf.append(ttf_test[0].split('url(')[1].split(');}<')[0])
            url='https://www.shixiseng.com'
            urllib.request.urlretrieve(url+ttf[0], "shixi.ttf")
        elif os.path.exists('shixi.ttf'):
            pass
        #获得ttf文件并解码
        else:
            ttf = []
            ttf_test=[]
            pwd = browser.find_elements_by_xpath('/html/head/style[1]')
            for x in pwd:
                ttf_test=[x.get_attribute('outerHTML')]
                ttf.append(ttf_test[0].split('url(')[1].split(');}<')[0])
            url='https://www.shixiseng.com'
            urllib.request.urlretrieve(url+ttf[0], "shixi.ttf")
        font = TTFont('shixi.ttf')
        # 由于实习增的这个文件页面刷新前后是不变的，所以不用前后进行字体文件的比对了
        font_base_order = font.getGlyphOrder()[2:]# 下载下来的文件头两个是空的
        # 新下载的问件与原文件进行比对
        # 前10个是0到9，从本地将对应的文字写出来
        map_list =[
            *[str(i) for i in range(10)], u'一', u'师', 'X', u'会', u'四', u'计', u'财', u'场', 'D', 'H',
            'L', 'P', 'T', u'聘', u'招', u'工', 'd', u'周', 'I', u'端', 'p', u'年', 'h', 'x', u'设', u'程',
            u'二', u'五', u'天', 't', 'C', 'G', u'前', 'K', 'O', u'网', 'S', 'W', 'c', 'g', 'k', 'o', 's',
            'w', u'广', u'市', u'月', u'个', 'B', 'F', u'告', 'N', 'R', 'V', 'Z', u'作', 'b', 'f', 'j', 'n',
            'r', 'v', 'z', u'三', u'互', u'生', u'人', u'政', 'A', 'J', 'E', 'I', u'件', 'M', '行', 'Q', 'U',
            'Y', 'a', 'e', 'i', 'm', u'软', 'q', 'u', u'银', 'y', u'联', 
        ]
        # 你会发现网页中编码对应的是font.getBestCmap()的key的16进制的值
        map_dict = {value: '&#' + hex(key)[1:]
                    for key, value in font.getBestCmap().items()}
        # 将固定的字体顺序和uni编码进行一一对应，并从map_dict中寻找16进制的值对应的字体
        temp_dict = {map_dict[key]: value for key, value in zip(font_base_order, map_list)}
        
        #地区
        dq=[]
        dq_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="__layout"]/div/div[2]/div[2]/div[1]/div[1]/div[1]/div/div[1]/div[1]/p[2]/span[1]')
        for x in pwd:
            dq_test=[x.text,x.get_attribute('href')]
            dq.append(dq_test[0])
        #公司名称
        gs=[]
        gs_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="__layout"]/div/div[2]/div[2]/div[1]/div[1]/div[1]/div/div[1]/div[2]/p[1]/a')
        for x in pwd:
            gs_test=[x.text,x.get_attribute('href')]
            gs.append(gs_test[0])
        #职位
        zw=[]
        zw_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="__layout"]/div/div[2]/div[2]/div[1]/div[1]/div[1]/div/div[1]/div[1]/p[1]/a')
        for x in pwd:
            zw_test=[x.get_attribute('outerHTML')]
            m=str(zw_test[0]).split('title="')[1].split('" target=')[0]
            m=''.join(m.split('amp;'))
            for k,value in temp_dict.items():
                m = m.replace(k,value)
            zw.append(m)
        #链接
        xq = []
        xq_test=[]
        pwd = browser.find_elements_by_xpath('//*[@id="__layout"]/div/div[2]/div[2]/div[1]/div[1]/div[1]/div/div[1]/div[1]/p[1]/a')
        for x in pwd:
            xq_test=[x.get_attribute("outerHTML")]
            xq.append(re.findall(r'href="(.*?)" title="',str(xq_test[0]))[0])
        if len(dq) != len(zw) or len(dq) != len(gs) or  len(gs) != len(zw) or len(xq) != len(zw):
            return []
        else:
            return [[dq[i],gs[i],zw[i],xq[i],'实习僧'] for i in range(len(xq))]  
    #实习信息的主函数  
    def shixi_main(self):
        #这里是全国实习信息版块
        path = os.path.abspath(os.curdir)
        sys.path.append(str(path))
        option = webdriver.ChromeOptions()
        option.add_argument('headless')
        browser = webdriver.Chrome(r'D:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe', options=option)
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        result_yingjiesheng=self.get_info_yingjiesheng()
        if result_yingjiesheng==[]:
            result_yingjiesheng=self.getinfo_new4(r'http://www.yingjiesheng.com')
        if result_yingjiesheng==[]:
            print('扎心了!应届生求职网的信息还没有更新......\n')
        url_quanguo = "https://www.shixiseng.com/interns?page={}&keyword=&type=intern&area=&months=&days=&degree=&official=&enterprise=IT300&salary=-0&publishTime=day&sortType=zj&city=全国&internExtend="
        url_wuhan = "https://www.shixiseng.com/interns?page=1&keyword=&type=intern&area=&months=&days=&degree=&official=&enterprise=&salary=-0&publishTime=day&sortType=zj&city=武汉&internExtend="
        result_shixisen = self.shuju_shixisen(1,browser,url_quanguo)
        for i in range(6):
            result_shixisen = self.zhenghe(result_shixisen, self.shuju_shixisen(i+2,browser,url_quanguo))

        #这里是武汉实习信息版块
        browser.get(url_wuhan)
        result_shixisen_wuhan = self.shuju_shixisen(1,browser,'no_url')
        for i in range(10):
            try:
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                button=browser.find_element_by_xpath('//*[@id="__layout"]/div/div[2]/div[2]/div[1]/div[1]/div[2]/div/ul/li[{}]'.format(i+2))
                button.click()
                time.sleep(1)
                result_shixisen_wuhan = self.zhenghe(result_shixisen_wuhan, self.shuju_shixisen(i+2,browser,'no_url'))
            except:
                break

        #信息筛选
        result_wuhan = [x for x in result_shixisen_wuhan if '武汉' in x[0]]
        browser.quit()
        #写入文件
        if result_yingjiesheng==[]:
            result = result_shixisen
        else:
            result = self.zhenghe(result_yingjiesheng,result_shixisen)
        if len(result) ==0:
            print('诶呀，实习僧网站似乎没有更新数据！\n')
        else:
            print('信息已经检索完毕，现在开始写入......\n')
            self.write_Excel_shixi(result,result_wuhan)
            print('信息已经写入完毕!\n')
            print('===============\n')    
        #print(result,result_shixisen_wuhan)
       

    #——————————————————————————————————————————————————————————————————————————全部信息的收取    
    def write_collection(self,result_jiuye,result_shixisen,result_wuhan,address):
        #写入信息
        print('信息搜索成功,开始写入....\n')
        wb = Workbook()
        re_now=[]
        re_next=[]
        d_now=datetime.datetime.now()
        d_now=d_now+datetime.timedelta(1)
        d_next=d_now+datetime.timedelta(1)
        for x in result_jiuye:
            if x[2][5:10]==d_now.strftime('%m-%d'):
                re_now.append(x)
            elif x[2][5:10]==d_next.strftime('%m-%d'):
                re_next.append(x)
        wb = Workbook()
        sheet = wb.active
        #sheet.title = str(d_now.strftime('%m-%d'))+'和'+str(d_next.strftime('%m-%d'))+'就业宣讲会信息汇总' # 创建最终保存表格
        sheet.title =str(d_now.strftime('%m'))+'月'+str(d_now.strftime('%d'))+'日-'+str(d_next.strftime('%m'))+'月'+str(d_next.strftime('%d'))+'日'+'就业宣讲会信息汇总'
        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 30
        #第一天
        sheet.cell(row=1, column=1, value=str(d_now.strftime('%m'))+'月'+str(d_now.strftime('%d'))+'日'+'就业宣讲会信息')
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        self.write_Excel_jiuye(re_now,2,wb)
        #第二天
        sheet.cell(row=len(re_now)+3, column=1, value=str(str(d_next.strftime('%m'))+'月'+str(d_next.strftime('%d'))+'日'+'就业宣讲会信息'))
        sheet.cell(row=len(re_now)+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=len(re_now)+3, start_column=1, end_row=len(re_now)+3, end_column=5)
        sheet.cell(row=len(re_now)+len(re_next)+5, column=1, value='武大研会发展权益部')
        sheet.merge_cells('A{}:E{}'.format(len(re_now)+len(re_next)+5,len(re_now)+len(re_next)+5))
        sheet['A{}'.format(len(re_now)+len(re_next)+5)].alignment = Alignment(horizontal='center', vertical='center')
        self.write_Excel_jiuye(re_next,len(re_now)+4,wb)
        #接下来写入实习信息
        sheet2 = wb.create_sheet(datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"实习信息" )
        sheet2.column_dimensions['A'].width = 9.0
        sheet2.column_dimensions['B'].width = 50.0
        sheet2.column_dimensions['C'].width = 50.0
        sheet2.column_dimensions['D'].width = 13.0
        sheet2.title = datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"实习信息" 
        sheet2.cell(row=1, column=1, value=datetime.datetime.now().strftime('%Y')+'年'+datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"实习信息")
        sheet2.merge_cells('A1:D1')
        #sheet.cellstyle('A1', font, align)
        sheet2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        # 创建最终保存表格
        sheet2.cell(row=2, column=1, value='地区')
        sheet2.cell(row=2, column=2, value='公司名称')
        sheet2.cell(row=2, column=3, value='岗位')
        sheet2.cell(row=2, column=4, value='链接')
        sheet.cell(row=2, column=5, value='来源')
        for i in range(5):
            for j in range(2,len(result_shixisen)+3):
                sheet2.cell(row=j, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
        
        for i in range(len(result_shixisen)):
            for j in range(len(result_shixisen[i])):
                if(j == 3):
                    sheet2.cell(row=i + 3, column=j+1, value=('=HYPERLINK("%s","%s")' % (result_shixisen[i][j], '详情')))
                else:
                    sheet2.cell(row=i + 3, column=j+1, value=str(result_shixisen[i][j]))
        #sheet['A3:D{}'.format(len(data)+2)].alignment = Alignment(horizontal='center', vertical='center')
        sheet2.cell(row=len(result_shixisen)+3, column=1, value='武大研会发展权益部')
        sheet2.merge_cells('A{}:D{}'.format(len(result_shixisen)+3,len(result_shixisen)+3))
        sheet2['A{}'.format(len(result_shixisen)+3)].alignment = Alignment(horizontal='center', vertical='center')
        #开启第二个sheet
        if result_wuhan == []:
            pass
        else:
            #首先要去除重复
            #result_shixisen_wuhan = list(set(result_shixisen_wuhan))
            list2=[]
            for i in result_wuhan:
                if i not in list2:
                    list2.append(i)
            result_wuhan=list2
            sheet1 = wb.create_sheet(datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"武汉地区实习信息" )
            sheet1.column_dimensions['A'].width = 9.0
            sheet1.column_dimensions['B'].width = 50.0
            sheet1.column_dimensions['C'].width = 50.0
            sheet1.column_dimensions['D'].width = 13.0
            #sheet1.title = datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"实习信息" 
            sheet1.cell(row=1, column=1, value=datetime.datetime.now().strftime('%Y')+'年'+datetime.datetime.now().strftime('%m')+'月'+datetime.datetime.now().strftime('%d')+'日'+"武汉地区实习信息")
            sheet1.merge_cells('A1:D1')
            #sheet.cellstyle('A1', font, align)
            sheet1['A1'].alignment = Alignment(horizontal='center', vertical='center')
            # 创建最终保存表格
            sheet1.cell(row=2, column=1, value='地区')
            sheet1.cell(row=2, column=2, value='公司名称')
            sheet1.cell(row=2, column=3, value='岗位')
            sheet1.cell(row=2, column=4, value='链接')
            sheet1.cell(row=2, column=5, value='来源')
            for i in range(5):
                for j in range(2,len(result_wuhan)+3):
                    sheet1.cell(row=j, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
            
            for i in range(len(result_wuhan)):
                for j in range(len(result_wuhan[i])):
                    if(j == 3):
                        sheet1.cell(row=i + 3, column=j+1, value=('=HYPERLINK("%s","%s")' % (result_wuhan[i][j], '详情')))
                    else:
                        sheet1.cell(row=i + 3, column=j+1, value=str(result_wuhan[i][j]))
            #sheet['A3:D{}'.format(len(data)+2)].alignment = Alignment(horizontal='center', vertical='center')
            sheet1.cell(row=len(result_wuhan)+3, column=1, value='武大研会发展权益部')
            sheet1.merge_cells('A{}:D{}'.format(len(result_wuhan)+3,len(result_wuhan)+3))
            sheet1['A{}'.format(len(result_wuhan)+3)].alignment = Alignment(horizontal='center', vertical='center')
        wb.save(address+'{}就业信息汇总.xlsx'.format(datetime.datetime.now().strftime('%m-%d')))        
    
    def collection(self,address):
        #打开浏览器
        print('就知道你会点这个，哈哈......\n')
        path = os.path.abspath(os.curdir)
        sys.path.append(str(path))
        option = webdriver.ChromeOptions()
        option.add_argument('headless')
        browser = webdriver.Chrome(r'D:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe', options=option)
        #browser.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        # 首先获取全部的实习信息
        result_yingjiesheng=self.get_info_yingjiesheng()
        if result_yingjiesheng==[]:
            result_yingjiesheng=self.getinfo_new4(r'http://www.yingjiesheng.com')
        if result_yingjiesheng==[]:
            print('扎心了!应届生求职网的信息还没有更新......\n')
        #开始抓取实习僧的信息
        url_quanguo = "https://www.shixiseng.com/interns?page={}&keyword=&type=intern&area=&months=&days=&degree=&official=&enterprise=IT300&salary=-0&publishTime=day&sortType=zj&city=全国&internExtend="
        url_wuhan = "https://www.shixiseng.com/interns?page=1&keyword=&type=intern&area=&months=&days=&degree=&official=&enterprise=&salary=-0&publishTime=day&sortType=zj&city=武汉&internExtend=1"
        #全国信息获取
        result_shixisen = self.shuju_shixisen(1,browser,url_quanguo)
        for i in range(5):
            result_shixisen = self.zhenghe(result_shixisen, self.shuju_shixisen(i+2,browser,url_quanguo))
        #整合应届生和实习僧的信息
        if result_yingjiesheng!=[]:
            result_shixisen = self.zhenghe(result_yingjiesheng, result_shixisen)
        #武汉实习信息版块
        browser.get(url_wuhan)
        result_shixisen_wuhan = self.shuju_shixisen(1,browser,'no_url')
        for i in range(10):
            try:
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                button=browser.find_element_by_xpath('//*[@id="__layout"]/div/div[2]/div[2]/div[1]/div[1]/div[2]/div/ul/li[{}]'.format(i+2))
                button.click()
                time.sleep(1)
                result_shixisen_wuhan = self.zhenghe(result_shixisen_wuhan, self.shuju_shixisen(i+2,browser,'no_url'))
            except:
                break
        #信息筛选
        result_wuhan = [x for x in result_shixisen_wuhan if '武汉' in x[0]]
        
        
        #之后获取全部的宣讲会信息
        print('开始抓取宣讲会信息......\n')
        #这里通过get请求需要模拟登录的页面
        browser.get("http://www.xsjy.whu.edu.cn/zftal-web/zfjy!wzxx/xjhxx_cxXjhForWeb.html")
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight)")
        #browser.manage().timeouts().implicitlyWait(2000, TimeUnit.SECONDS);#延迟
        k = 3
        if k==1:
            result_jiuye = self.shuju_jiuye(browser)
        else: 
            result_jiuye = self.shuju_jiuye(browser)
            for i in range(k-1):
                button = WebDriverWait(browser,10).until(EC.element_to_be_clickable((By.LINK_TEXT,"下一页")))
                button.click()  
                result_jiuye += self.shuju_jiuye(browser)
        if result_jiuye ==[]:
            print('网站似乎没有更新数据，可能是明后两天没有宣讲会......\n')
        browser.quit()
        self.write_collection(result_jiuye,result_shixisen,result_wuhan,address)
   

#—————————————————————————————————空中宣讲会—————————————————————————————————————————
    def write_Excel_kongzhong(self,data):
        #数据初始化
        re_now=[]
        re_next=[]
        re_next2=[]
        d_now=datetime.datetime.now()
        d_now=d_now+datetime.timedelta(1)
        d_next=d_now+datetime.timedelta(1)
        d_next2=d_next+datetime.timedelta(1)
        for x in data:
            if len(re_now)==0 and x[1][2:10]==d_now.strftime('%y-%m-%d'):
                re_now.append(x)
            elif len(re_now)!=0 and x[1][2:10]==d_now.strftime('%y-%m-%d') and  x[2] not in [y[2] for y in re_now]:
                re_now.append(x)           
            elif len(re_next)==0 and x[1][2:10]==d_next.strftime('%y-%m-%d'):
                re_next.append(x)
            elif len(re_next)!=0 and x[1][2:10]==d_next.strftime('%y-%m-%d') and  x[2] not in [y[2] for y in re_next]:
                re_next.append(x)
            elif len(re_next2)==0 and x[1][2:10]==d_next2.strftime('%y-%m-%d'):
                re_next2.append(x)
            elif len(re_next2)!=0 and x[1][2:10]==d_next2.strftime('%y-%m-%d') and  x[2] not in [y[2] for y in re_next2]:
                re_next2.append(x)
        #开始写入----初始化
        wb = Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 6.0
        sheet.column_dimensions['B'].width = 17.0
        sheet.column_dimensions['C'].width = 50.0
        sheet.column_dimensions['D'].width = 30.0
        sheet.column_dimensions['E'].width = 50.0
        sheet.column_dimensions['F'].width = 6.0
        #标题
        title_day=str(d_now.strftime('%m'))+'月'+str(d_now.strftime('%d'))+'日-'+str(d_next2.strftime('%m'))+'月'+str(d_next2.strftime('%d'))+'日'+'武汉地区空中宣讲会信息汇总'
        sheet.title =title_day
        #第一天的题头
        sheet.cell(row=1, column=1, value=str(d_now.strftime('%m'))+'月'+str(d_now.strftime('%d'))+'日'+'武汉地区空中宣讲会信息')
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        sheet.cell(row=2, column=1, value='地区')
        sheet.cell(row=2, column=2, value='时间')
        sheet.cell(row=2, column=3, value='公司名称')
        sheet.cell(row=2, column=4, value='举办地点')
        sheet.cell(row=2, column=5, value='类型')
        sheet.cell(row=2, column=6, value='链接')
        #第一天数据
        for i in range(len(re_now)):
            for j in range(1, len(re_now[i])+1):
                if(j == len(re_now[i])):
                    sheet.cell(row=i + 3, column=6, value=('=HYPERLINK("%s","%s")' % (re_now[i][-1], '详情')))
                else:
                    sheet.cell(row=i + 3, column=j, value=str(re_now[i][j-1]))
        #di一天输入完毕-开始第二天
        sheet.cell(row=len(re_now)+3, column=1, value=str(d_next.strftime('%m'))+'月'+str(d_next.strftime('%d'))+'日'+'武汉地区空中宣讲会信息')
        sheet.cell(row=len(re_now)+3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=len(re_now)+3, start_column=1, end_row=len(re_now)+3, end_column=6)
        sheet.cell(row=len(re_now)+4, column=1, value='地区')
        sheet.cell(row=len(re_now)+4, column=2, value='时间')
        sheet.cell(row=len(re_now)+4, column=3, value='公司名称')
        sheet.cell(row=len(re_now)+4, column=4, value='举办地点')
        sheet.cell(row=len(re_now)+4, column=5, value='类型')
        sheet.cell(row=len(re_now)+4, column=6, value='链接')
        #第2天数据
        for i in range(len(re_next)):
            for j in range(1, len(re_next[i])+1):
                if(j == len(re_next[i])):
                    sheet.cell(row=i + len(re_now)+5, column=6, value=('=HYPERLINK("%s","%s")' % (re_next[i][-1], '详情')))
                else:
                    sheet.cell(row=i + len(re_now)+5, column=j, value=str(re_next[i][j-1]))
        #2天输入完毕-开始第3天
        sheet.cell(row=len(re_now)+len(re_next)+5, column=1, value=str(d_next2.strftime('%m'))+'月'+str(d_next2.strftime('%d'))+'日'+'武汉地区空中宣讲会信息')
        sheet.cell(row=len(re_now)+len(re_next)+5, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row=len(re_now)+len(re_next)+5, start_column=1, end_row=len(re_now)+len(re_next)+5, end_column=6)
        sheet.cell(row=len(re_now)+len(re_next)+6, column=1, value='地区')
        sheet.cell(row=len(re_now)+len(re_next)+6, column=2, value='时间')
        sheet.cell(row=len(re_now)+len(re_next)+6, column=3, value='公司名称')
        sheet.cell(row=len(re_now)+len(re_next)+6, column=4, value='举办地点')
        sheet.cell(row=len(re_now)+len(re_next)+6, column=5, value='类型')
        sheet.cell(row=len(re_now)+len(re_next)+6, column=6, value='链接')
        #第3天数据
        for i in range(len(re_next2)):
            for j in range(1, len(re_next2[i])+1):
                if(j == len(re_next2[i])):
                    sheet.cell(row=i + len(re_now)+len(re_next)+7, column=6, value=('=HYPERLINK("%s","%s")' % (re_next2[i][-1], '详情')))
                else:
                    sheet.cell(row=i + len(re_now)+len(re_next)+7, column=j, value=str(re_next2[i][j-1]))
        #标志
        sheet.cell(row=len(re_now)+len(re_next)+len(re_next2)+7, column=1, value='武大研会发展权益部')
        sheet.merge_cells('A{}:F{}'.format(len(re_now)+len(re_next)+len(re_next2)+7,len(re_now)+len(re_next)+len(re_next2)+7))
        sheet['A{}'.format(len(re_now)+len(re_next)+len(re_next2)+7)].font = Font('黑体', bold=True)
        sheet['A{}'.format(len(re_now)+len(re_next)+len(re_next2)+7)].alignment = Alignment(horizontal='center', vertical='center')
                    
        #save
        try:
            wb.save('D:/发展权益部/'+title_day+'.xlsx')
            print('文件保存在: D:/发展权益部/'+title_day+'.xlsx')
        except FileNotFoundError:
            try:
                os.makedirs(r"D:/发展权益部")
                wb.save(r'D:/发展权益部/'+title_day+'.xlsx')
                print('文件保存在: D:/发展权益部/'+title_day+'.xlsx')
            except FileNotFoundError:
                os.makedirs(r"C:/发展权益部")
                wb.save(r'C:/发展权益部/'+title_day+'.xlsx')
                print('文件保存在: C:/发展权益部/'+title_day+'.xlsx')
        return title_day        
        
    def get_result(self,browser,url):
        
        browser.get(url)
        #第一页的数据
        #获取宣讲会名称mc
        mc=[]
        mc_test=[]
        pwd = browser.find_elements_by_xpath('/html/body/div[5]/div[2]/form/div[1]/table')
        #//*[@id="wide"]/div[5]/form/div[2]
        for x in pwd:
            mc_test=[x.text]
            mc.append(mc_test[0])
            print(mc_test[0])
        
        total=[x.split(' ') for x in mc[0].split('\n')]
        xq=[]
        xq_test=[]
        pwd = browser.find_elements_by_xpath('//html/body/div[5]/div[2]/form/div[1]/table')
        #//*[@id="wide"]/div[5]/form/div[2]
        for x in pwd:
            xq_test=[x.get_attribute('innerHTML')]
            xq.append(xq_test[0])
        url=re.findall(r'href="(.*?)" target',str(xq_test[0]))
        url_true=[url[i] for i in range(len(url)) if(i+1)%4==0]
        for i in range(len(url_true)):
            if 'http' not in str(url_true[i]):
                url_true[i]='http://my.yingjiesheng.com'+url_true[i]
        result=[total[i]+[url_true[i]] for i in range(len(url_true))]
        # print(result)
        for i in range(len(result)):
            if result[i][-3]=='已取消' or result[i][-3]=='取消':
                result[i]=[]
        result=[x for x in result if x!=[]]
        return result        
    
    
    
    def kongzhong(self):
        global title
        print('开始抓取空中宣讲会信息')
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')
        browser = webdriver.Chrome(r'D:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe',options=chrome_options)
        result=[]
        for i in range(7):
            url = r"http://my.yingjiesheng.com/index.php/personal/xjhinfo.htm/?page={}&cid=&city=19&word=&province=0&schoolid=&sdate=&hyid=0".format(i+1)
            result+=self.get_result(browser,url)
        browser.quit()
    #     print(result)
        for i in range(len(result)):
            if len(result[i])!=7:
                result[i]=result[i][:4]+[result[i][4]+' '+result[i][5]]+result[i][-2:]
        if result ==[]:
            print('网站似乎没有更新数据，请稍后再来,也有可能是明后两天没有宣讲会！')
        else:
            title = self.write_Excel_kongzhong(result)
        
#—————————————————————————————————打开文件机制—————————————————————————————————————————
    def open( self,open_pwd):
        global title
        if open_pwd==1:
            command =r"D:/发展权益部/{}实习信息.xlsx".format(datetime.datetime.now().strftime('%m-%d'))
            os.system(command)
            command2 ="C:/发展权益部/{}实习信息.xlsx".format(datetime.datetime.now().strftime('%m-%d'))
            os.system(command2)
        if open_pwd==2:
            command ="D:/发展权益部/明后就业宣讲会信息.xlsx"
            os.system(command)
            command2 ="C:/发展权益部/明后就业宣讲会信息.xlsx"
            os.system(command2)
        if open_pwd==3:
            command ="D:/发展权益部/{}就业信息汇总.xlsx".format(datetime.datetime.now().strftime('%m-%d'))
            os.system(command)
            command2 ="C:/发展权益部/{}就业信息汇总.xlsx".format(datetime.datetime.now().strftime('%m-%d'))
            os.system(command2)
        if open_pwd==4:
            command ='D:/发展权益部/'+title+'.xlsx'
            os.system(command)
            command2 ='C:/发展权益部/'+title+'.xlsx'
            os.system(command2)
    def main(self):
        index = input('请输入所需要的信息序号：')
        if index=='1':
            self.shixi_main()
            self.open(1)
            index2= input('是否需要其他的信息？（是\否）：')
            if index2 =='是':
                self.menu()
                self.main()
            else:
                print('感谢使用')
        elif index == '2':
            self.jiuye_main()
            self.open(2)
            index2= input('是否需要其他的信息？（是\否）：')
            if index2 =='是':
                self.menu()
                self.main()
            else:
                print('感谢使用')
        elif index == '3':
            try:
                address='D:/发展权益部/'
                self.collection(address)
            except FileNotFoundError:
                address='C:/发展权益部/'
                self.collection(address)
            self.open(3)  
            index2= input('是否需要其他的信息？（是\否）：')
            if index2 =='是':
                self.menu()
                self.main()
            else:
                print('感谢使用')
        elif index == '4':
            self.kongzhong()
            self.open(4)
            index2= input('是否需要其他的信息？（是\否）：')
            if index2 =='是':
                self.menu()
                self.main()
            else:
                print('感谢使用')
        else:
            print('请输入正确的信息！')
            self.main()
    def menu(self):
        print('======================================================================')
        print('	    如果需要【实  习  信  息】，请输入 1 之后回车')
        print('	    如果需要【就业宣讲会信息】，请输入 2 之后回车')
        print('	    如果需要【全  部  都  要】，请输入 3 之后回车')
        print('	    如果需要【空中宣讲会信息】，请输入 4 之后回车') 
        print('======================================================================')      
#main       
if __name__ == '__main__': 
    print('----------------------------------------------------------------------')
    print('---------欢迎来到武汉大学研究生会发展权益部的就业信息站更新系统-----------')
    print('----------------------------------------------------------------------')
    print('使用之前请特别注意：')
    print('	“就业宣讲会信息”是指武汉大学求职信息网之中的宣讲会信息!')
    print('    “空中宣讲会信息”是指武汉地区的空中宣讲会，小伙伴们千万不要弄混了！')
    print('		疫情期间，推荐使用功能 1 和功能 4 ！')
    print('----------------------------------------------------------------------')
    print('		请选择您所需要的信息，信息选项如下：')
    print('======================================================================')
    print('	    如果需要【实  习  信  息】，请输入 1 之后回车')
    print('	    如果需要【就业宣讲会信息】，请输入 2 之后回车')
    print('	    如果需要【全  部  都  要】，请输入 3 之后回车')
    print('	    如果需要【空中宣讲会信息】，请输入 4 之后回车') 
    print('======================================================================')      
   

           # 初始化
    start = main()  # 实例MyFrame类，并传递参数  